from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction, models
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Sum, F
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from .models import (
    Warehouse, MaterialCategory, ProductCategory,
    RawMaterial, FinishedProduct, InventoryTransaction, StockAlert
)
from .forms import (
    WarehouseForm, MaterialCategoryForm, ProductCategoryForm,
    RawMaterialForm, FinishedProductForm, InventoryTransactionForm, StockAdjustmentForm
)
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction, models
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, Sum, F
from django.utils import timezone
from .models import (
    Warehouse, MaterialCategory, ProductCategory,
    RawMaterial, FinishedProduct, InventoryTransaction, StockAlert
)
from .forms import (
    WarehouseForm, MaterialCategoryForm, ProductCategoryForm,
    RawMaterialForm, FinishedProductForm, InventoryTransactionForm, StockAdjustmentForm
)


@login_required
def inventory_list(request):
    """Main inventory dashboard view"""
    # Get summary statistics
    raw_materials_count = RawMaterial.objects.filter(is_active=True).count()
    finished_products_count = FinishedProduct.objects.filter(is_active=True).count()
    warehouses_count = Warehouse.objects.filter(is_active=True).count()

    # Get low stock alerts
    low_stock_raw = RawMaterial.objects.filter(
        is_active=True,
        current_stock__lte=F('minimum_stock')
    ).count()
    low_stock_finished = FinishedProduct.objects.filter(
        is_active=True,
        current_stock__lte=F('minimum_stock')
    ).count()
    total_low_stock = low_stock_raw + low_stock_finished

    # Get recent transactions (last 5)
    recent_transactions = InventoryTransaction.objects.all().order_by('-created_at')[:5]

    # Get low stock items (last 5)
    low_stock_items = []
    low_stock_raw_items = RawMaterial.objects.filter(
        is_active=True,
        current_stock__lte=F('minimum_stock')
    ).order_by('current_stock')[:5]
    low_stock_finished_items = FinishedProduct.objects.filter(
        is_active=True,
        current_stock__lte=F('minimum_stock')
    ).order_by('current_stock')[:5]

    # Combine and sort by stock level
    for item in low_stock_raw_items:
        low_stock_items.append({
            'name': item.name,
            'code': item.code,
            'current_stock': item.current_stock,
            'minimum_stock': item.minimum_stock,
            'type': 'Raw Material'
        })
    for item in low_stock_finished_items:
        low_stock_items.append({
            'name': item.name,
            'code': item.code,
            'current_stock': item.current_stock,
            'minimum_stock': item.minimum_stock,
            'type': 'Finished Product'
        })

    # Sort by current stock (lowest first)
    low_stock_items.sort(key=lambda x: x['current_stock'])

    context = {
        'raw_materials_count': raw_materials_count,
        'finished_products_count': finished_products_count,
        'warehouses_count': warehouses_count,
        'total_low_stock': total_low_stock,
        'recent_transactions': recent_transactions,
        'low_stock_items': low_stock_items[:5],  # Show only first 5
    }
    return render(request, 'inventory/dashboard.html', context)


# Warehouse Views
@login_required
def warehouse_list(request):
    warehouses = Warehouse.objects.all().order_by('-created_at')
    paginator = Paginator(warehouses, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'title': 'Warehouses'
    }
    return render(request, 'inventory/warehouse_list.html', context)


@login_required
def warehouse_create(request):
    if request.method == 'POST':
        form = WarehouseForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Warehouse created successfully.')
            return redirect('warehouse_list')
    else:
        form = WarehouseForm()

    context = {
        'form': form,
        'title': 'Create Warehouse'
    }
    return render(request, 'inventory/warehouse_form.html', context)


@login_required
def warehouse_update(request, pk):
    warehouse = get_object_or_404(Warehouse, pk=pk)
    if request.method == 'POST':
        form = WarehouseForm(request.POST, instance=warehouse)
        if form.is_valid():
            form.save()
            messages.success(request, 'Warehouse updated successfully.')
            return redirect('warehouse_list')
    else:
        form = WarehouseForm(instance=warehouse)

    context = {
        'form': form,
        'warehouse': warehouse,
        'title': 'Update Warehouse'
    }
    return render(request, 'inventory/warehouse_form.html', context)


@login_required
def warehouse_delete(request, pk):
    warehouse = get_object_or_404(Warehouse, pk=pk)
    if request.method == 'POST':
        warehouse.delete()
        messages.success(request, 'Warehouse deleted successfully.')
        return redirect('warehouse_list')

    context = {
        'warehouse': warehouse,
        'title': 'Delete Warehouse'
    }
    return render(request, 'inventory/warehouse_confirm_delete.html', context)


# Raw Material Views
@login_required
def raw_material_list(request):
    raw_materials = RawMaterial.objects.all().order_by('code')
    paginator = Paginator(raw_materials, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'title': 'Raw Materials'
    }
    return render(request, 'inventory/raw_material_list.html', context)


@login_required
def raw_material_create(request):
    if request.method == 'POST':
        form = RawMaterialForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Raw material created successfully.')
            return redirect('raw_material_list')
    else:
        form = RawMaterialForm()

    context = {
        'form': form,
        'title': 'Create Raw Material'
    }
    return render(request, 'inventory/raw_material_form.html', context)


@login_required
def raw_material_update(request, pk):
    raw_material = get_object_or_404(RawMaterial, pk=pk)
    if request.method == 'POST':
        form = RawMaterialForm(request.POST, instance=raw_material)
        if form.is_valid():
            form.save()
            messages.success(request, 'Raw material updated successfully.')
            return redirect('raw_material_list')
    else:
        form = RawMaterialForm(instance=raw_material)

    context = {
        'form': form,
        'raw_material': raw_material,
        'title': 'Update Raw Material'
    }
    return render(request, 'inventory/raw_material_form.html', context)


@login_required
def raw_material_delete(request, pk):
    raw_material = get_object_or_404(RawMaterial, pk=pk)
    if request.method == 'POST':
        raw_material.delete()
        messages.success(request, 'Raw material deleted successfully.')
        return redirect('raw_material_list')

    context = {
        'raw_material': raw_material,
        'title': 'Delete Raw Material'
    }
    return render(request, 'inventory/raw_material_confirm_delete.html', context)


# Finished Product Views
@login_required
def finished_product_list(request):
    finished_products = FinishedProduct.objects.all().order_by('code')
    paginator = Paginator(finished_products, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'title': 'Finished Products'
    }
    return render(request, 'inventory/finished_product_list.html', context)


@login_required
def finished_product_create(request):
    if request.method == 'POST':
        form = FinishedProductForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Finished product created successfully.')
            return redirect('finished_product_list')
    else:
        form = FinishedProductForm()

    context = {
        'form': form,
        'title': 'Create Finished Product'
    }
    return render(request, 'inventory/finished_product_form.html', context)


@login_required
def finished_product_update(request, pk):
    finished_product = get_object_or_404(FinishedProduct, pk=pk)
    if request.method == 'POST':
        form = FinishedProductForm(request.POST, instance=finished_product)
        if form.is_valid():
            form.save()
            messages.success(request, 'Finished product updated successfully.')
            return redirect('finished_product_list')
    else:
        form = FinishedProductForm(instance=finished_product)

    context = {
        'form': form,
        'finished_product': finished_product,
        'title': 'Update Finished Product'
    }
    return render(request, 'inventory/finished_product_form.html', context)


@login_required
def finished_product_delete(request, pk):
    finished_product = get_object_or_404(FinishedProduct, pk=pk)
    if request.method == 'POST':
        finished_product.delete()
        messages.success(request, 'Finished product deleted successfully.')
        return redirect('finished_product_list')

    context = {
        'finished_product': finished_product,
        'title': 'Delete Finished Product'
    }
    return render(request, 'inventory/finished_product_confirm_delete.html', context)


# Stock Adjustment Views
@login_required
def stock_adjustment(request):
    if request.method == 'POST':
        form = StockAdjustmentForm(request.POST)
        if form.is_valid():
            material_type = form.cleaned_data['material_type']
            material_id = form.cleaned_data['material_id']
            adjustment_type = form.cleaned_data['adjustment_type']
            quantity = form.cleaned_data['quantity']
            reason = form.cleaned_data['reason']
            warehouse = form.cleaned_data['warehouse']

            with transaction.atomic():
                if material_type == 'raw':
                    material = get_object_or_404(RawMaterial, pk=material_id)
                    if adjustment_type == 'add':
                        material.current_stock += quantity
                    else:
                        material.current_stock -= quantity
                    material.save()
                else:
                    material = get_object_or_404(FinishedProduct, pk=material_id)
                    if adjustment_type == 'add':
                        material.current_stock += quantity
                    else:
                        material.current_stock -= quantity
                    material.save()

                # Create transaction record
                transaction_type = 'IN' if adjustment_type == 'add' else 'ADJ'
                InventoryTransaction.objects.create(
                    transaction_type=transaction_type,
                    material_type=material_type,
                    material_id=material_id,
                    material_name=material.name,
                    quantity=quantity,
                    unit_price=material.unit_price,
                    reference_number=f"ADJ-{request.user.id}-{timezone.now().strftime('%Y%m%d%H%M%S')}",
                    notes=f"Stock adjustment: {reason}",
                    warehouse=warehouse,
                    created_by=request.user
                )

            messages.success(request, 'Stock adjustment completed successfully.')
            return redirect('stock_adjustment')
    else:
        form = StockAdjustmentForm()

    context = {
        'form': form,
        'title': 'Stock Adjustment'
    }
    return render(request, 'inventory/stock_adjustment.html', context)


@login_required
def get_material_details(request):
    """AJAX view to get material details for stock adjustment"""
    material_type = request.GET.get('material_type')
    material_id = request.GET.get('material_id')

    if material_type == 'raw':
        material = get_object_or_404(RawMaterial, pk=material_id)
    else:
        material = get_object_or_404(FinishedProduct, pk=material_id)

    data = {
        'name': material.name,
        'current_stock': str(material.current_stock),
        'unit': material.unit if hasattr(material, 'unit') else 'pcs'
    }
    return JsonResponse(data)


# Transaction History Views
@login_required
def transaction_list(request):
    transactions = InventoryTransaction.objects.all().order_by('-created_at')
    paginator = Paginator(transactions, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

@login_required
def download_raw_material_template(request):
    """Download XLSX template for raw material bulk import"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Raw Materials Template"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "Code", "Name", "Category", "Unit", "Description",
        "Minimum Stock", "Current Stock", "Unit Price"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Sample data
    sample_data = [
        ["RM001", "Leather", "Leather", "kg", "Premium leather material", 10, 50, 25.50],
        ["RM002", "Thread", "Thread", "roll", "Nylon thread", 5, 20, 5.00],
        ["RM003", "Sole Rubber", "Rubber", "kg", "Rubber for shoe soles", 15, 100, 12.75]
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Unit choices note
    ws.cell(row=len(sample_data) + 3, column=1, value="Unit choices: kg, m, pcs, roll, sheet")
    ws.cell(row=len(sample_data) + 4, column=1, value="Category: Must match existing MaterialCategory name")

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="raw_material_template.xlsx"'
    wb.save(response)
    return response


@login_required
def download_finished_product_template(request):
    """Download XLSX template for finished product bulk import"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Finished Products Template"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "Code", "Name", "Category", "Size", "Color", "Description",
        "Current Stock", "Minimum Stock", "Unit Price"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Sample data
    sample_data = [
        ["FP001", "Running Shoes", "Sports", "42", "black", "Comfortable running shoes", 25, 5, 89.99],
        ["FP002", "Casual Sneakers", "Casual", "38", "white", "Everyday wear sneakers", 30, 10, 65.50],
        ["FP003", "Formal Shoes", "Formal", "41", "brown", "Business formal shoes", 15, 3, 120.00]
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Choices notes
    ws.cell(row=len(sample_data) + 3, column=1, value="Size choices: 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45")
    ws.cell(row=len(sample_data) + 4, column=1, value="Color choices: black, white, brown, blue, red, green, yellow, gray, navy, beige")
    ws.cell(row=len(sample_data) + 5, column=1, value="Category: Must match existing ProductCategory name")

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="finished_product_template.xlsx"'
    wb.save(response)
    return response


@login_required
def bulk_import_raw_materials(request):
    """Bulk import raw materials from XLSX file"""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            success_count = 0
            error_messages = []
            row_num = 1  # Skip header

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_num += 1

                # Skip empty rows
                if not any(row):
                    continue

                try:
                    code, name, category_name, unit, description, min_stock, current_stock, unit_price = row[:8]

                    # Validate required fields
                    if not code or not name or not category_name:
                        error_messages.append(f"Row {row_num}: Code, Name, and Category are required")
                        continue

                    # Get or create category
                    category, created = MaterialCategory.objects.get_or_create(name=category_name.strip())

                    # Validate unit (case-insensitive)
                    valid_units = ['kg', 'm', 'pcs', 'roll', 'sheet']
                    unit_lower = unit.lower().strip() if unit else ''
                    if unit_lower not in valid_units:
                        error_messages.append(f"Row {row_num}: Invalid unit '{unit}'. Must be one of: {', '.join(valid_units)}")
                        continue

                    # Create raw material
                    raw_material, created = RawMaterial.objects.get_or_create(
                        code=code.strip(),
                        defaults={
                            'name': name.strip(),
                            'category': category,
                            'unit': unit_lower,
                            'description': description.strip() if description else '',
                            'minimum_stock': float(min_stock or 0),
                            'current_stock': float(current_stock or 0),
                            'unit_price': float(unit_price or 0),
                        }
                    )

                    if not created:
                        # Update existing material (add to stock instead of setting)
                        raw_material.name = name.strip()
                        raw_material.category = category
                        raw_material.unit = unit_lower
                        raw_material.description = description.strip() if description else ''
                        raw_material.minimum_stock = float(min_stock or 0)
                        raw_material.current_stock += float(current_stock or 0)  # Add to existing stock
                        raw_material.unit_price = float(unit_price or 0)
                        raw_material.save()

                    success_count += 1

                except Exception as e:
                    error_messages.append(f"Row {row_num}: {str(e)}")

            if success_count > 0:
                messages.success(request, f'Successfully imported {success_count} raw materials.')
            if error_messages:
                for error in error_messages[:10]:  # Show first 10 errors
                    messages.warning(request, error)
                if len(error_messages) > 10:
                    messages.warning(request, f'... and {len(error_messages) - 10} more errors.')

            return redirect('raw_material_list')

        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            return redirect('bulk_import_raw_materials')

    return render(request, 'inventory/bulk_import_raw_materials.html', {
        'title': 'Bulk Import Raw Materials'
    })


@login_required
def bulk_import_finished_products(request):
    """Bulk import finished products from XLSX file"""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            success_count = 0
            error_messages = []
            row_num = 1  # Skip header

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_num += 1

                # Skip empty rows
                if not any(row):
                    continue

                try:
                    code, name, category_name, size, color, description, current_stock, min_stock, unit_price = row[:9]

                    # Validate required fields
                    if not code or not name or not category_name or not size or not color:
                        error_messages.append(f"Row {row_num}: Code, Name, Category, Size, and Color are required")
                        continue

                    # Get or create category
                    category, created = ProductCategory.objects.get_or_create(name=category_name.strip())

                    # Validate size
                    valid_sizes = ['35', '36', '37', '38', '39', '40', '41', '42', '43', '44', '45']
                    if str(size) not in valid_sizes:
                        error_messages.append(f"Row {row_num}: Invalid size '{size}'. Must be one of: {', '.join(valid_sizes)}")
                        continue

                    # Validate color (case-insensitive)
                    valid_colors = ['black', 'white', 'brown', 'blue', 'red', 'green', 'yellow', 'gray', 'navy', 'beige', 'sands']
                    color_lower = color.lower().strip() if color else ''
                    if color_lower not in valid_colors:
                        error_messages.append(f"Row {row_num}: Invalid color '{color}'. Must be one of: {', '.join(valid_colors)}")
                        continue

                    # Create finished product
                    product, created = FinishedProduct.objects.get_or_create(
                        code=code.strip(),
                        defaults={
                            'name': name.strip(),
                            'category': category,
                            'size': size,
                            'color': color_lower,
                            'description': description.strip() if description else '',
                            'current_stock': int(current_stock or 0),
                            'minimum_stock': int(min_stock or 0),
                            'unit_price': float(unit_price or 0),
                        }
                    )

                    if not created:
                        # Update existing product (add to stock instead of setting)
                        product.name = name.strip()
                        product.category = category
                        product.size = size
                        product.color = color_lower
                        product.description = description.strip() if description else ''
                        product.current_stock += int(current_stock or 0)  # Add to existing stock
                        product.minimum_stock = int(min_stock or 0)
                        product.unit_price = float(unit_price or 0)
                        product.save()

                    success_count += 1

                except Exception as e:
                    error_messages.append(f"Row {row_num}: {str(e)}")

            if success_count > 0:
                messages.success(request, f'Successfully imported {success_count} finished products.')
            if error_messages:
                for error in error_messages[:10]:  # Show first 10 errors
                    messages.warning(request, error)
                if len(error_messages) > 10:
                    messages.warning(request, f'... and {len(error_messages) - 10} more errors.')

            return redirect('finished_product_list')

        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            return redirect('bulk_import_finished_products')

    return render(request, 'inventory/bulk_import_finished_products.html', {
        'title': 'Bulk Import Finished Products'
    })
    context = {
        'page_obj': page_obj,
        'title': 'Transaction History'
    }
    return render(request, 'inventory/transaction_list.html', context)


# Stock Alert Views
@login_required
def stock_alert_list(request):
    # Generate dynamic alerts from current stock levels
    alerts = []

    # Get low stock raw materials
    low_stock_raw = RawMaterial.objects.filter(
        is_active=True,
        current_stock__lte=F('minimum_stock')
    ).order_by('current_stock')

    for material in low_stock_raw:
        alert_type = 'out_of_stock' if material.current_stock <= 0 else 'low_stock'
        message = f"Stock level ({material.current_stock}) is below minimum ({material.minimum_stock})"

        alerts.append({
            'alert_type': alert_type,
            'material_type': 'raw',
            'material_id': material.id,
            'material_name': material.name,
            'current_stock': material.current_stock,
            'threshold': material.minimum_stock,
            'message': message,
            'created_at': material.updated_at,  # Use updated_at as created_at for dynamic alerts
        })

    # Get low stock finished products
    low_stock_finished = FinishedProduct.objects.filter(
        is_active=True,
        current_stock__lte=F('minimum_stock')
    ).order_by('current_stock')

    for product in low_stock_finished:
        alert_type = 'out_of_stock' if product.current_stock <= 0 else 'low_stock'
        message = f"Stock level ({product.current_stock}) is below minimum ({product.minimum_stock})"

        alerts.append({
            'alert_type': alert_type,
            'material_type': 'finished',
            'material_id': product.id,
            'material_name': product.name,
            'current_stock': product.current_stock,
            'threshold': product.minimum_stock,
            'message': message,
            'created_at': product.updated_at,  # Use updated_at as created_at for dynamic alerts
        })

    # Sort alerts by current stock (lowest first)
    alerts.sort(key=lambda x: x['current_stock'])

    # Paginate the alerts
    paginator = Paginator(alerts, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'title': 'Stock Alerts'
    }
    return render(request, 'inventory/stock_alert_list.html', context)
